import tkinter as tk
from tkinter import ttk, messagebox
import cv2
import threading
import time
import re
import unicodedata
from datetime import datetime, date
from PIL import Image, ImageTk
import openpyxl
from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
import os

# ─── OCR ENGINE ───────────────────────────────────────────────
try:
    import easyocr
    OCR_ENGINE = "easyocr"
    _reader = None
except ImportError:
    try:
        import pytesseract
        OCR_ENGINE = "tesseract"
    except ImportError:
        OCR_ENGINE = "none"

C = {
    'bg':       '#111318',
    'surface':  '#1C1F26',
    'surface2': '#242830',
    'border':   '#2E3340',
    'accent':   '#4F8EF7',
    'accent2':  '#38D9A9',
    'warn':     '#F5A623',
    'danger':   '#F05C5C',
    'text':     '#E8EAF0',
    'muted':    '#6B7280',
    'muted2':   '#9CA3AF',
    'success':  '#34D399',
    'panel':    '#161A21',
}

# ── BANDEIRAS ─────────────────────────────────────────────────
# Cada entrada: (regex_no_texto_completo, nome_normalizado)
# Testados na ordem da lista — primeiro match vence.
BANDEIRAS_REGRAS = [
    # ── PIX ──────────────────────────────────────────────────────
    (r'pagamento\s*p[il1]x|p[il1]x\s*bacen|\bp[il1]x\b', 'PIX'),

    # ── PRÉ-PAGOS (antes dos genéricos) ─────────────────────────
    # "PREPAGO MASTERCARD" ou "MASTERCARD PREPAGO" e variações com erro OCR
    (r'pre.{0,3}pago\s*m[ao4]ster|m[ao4]ster.{0,5}pre.{0,3}pago', 'Mastercard Pré-pago'),
    (r'pre.{0,3}pago\s*v[il1]sa|v[il1]sa\s*pre.{0,3}pago',        'Visa Pré-pago'),
    (r'pre.{0,3}pago\s*el[o0]|el[o0]\s*pre.{0,3}pago',            'Elo Pré-pago'),
    (r'cabal\s*pre.{0,3}pago|pre.{0,3}pago\s*cabal',              'Cabal Pré-pago'),

    # ── VISA ──────────────────────────────────────────────────────
    # OCR pode ler "V1SA", "VJSA", "VÏSA"
    (r'v[il1j][s5][a4]\s*electron',   'Visa Electron'),
    (r'v[il1j][s5][a4]\s*vale',       'Visa Vale'),
    (r'\bv[il1j][s5][a4]\b',          'Visa'),

    # ── MASTERCARD ───────────────────────────────────────────────
    # OCR pode ler "HASTERCARD", "MAST3RCARD", etc.
    (r'\bm[a4][s5]t[e3]r\s*c[a4]rd\b|\bm[a4][s5]t[e3]rc[a4]rd\b', 'Mastercard'),
    (r'\bm[a4][s5]t[e3]r\b',         'Mastercard'),
    (r'\bm[a4][e3]stro\b',           'Maestro'),

    # ── ELO ──────────────────────────────────────────────────────
    # OCR pode ler "EL0", "ELQ"
    (r'\bel[o0q]\b',                  'Elo'),

    # ── HIPERCARD ────────────────────────────────────────────────
    (r'h[il1]p[e3]r\s*c[a4]rd|h[il1]p[e3]rc[a4]rd|\bh[il1]p[e3]r\b', 'Hipercard'),

    # ── AMERICAN EXPRESS ─────────────────────────────────────────
    (r'am[e3]r[il1]c[a4]n\s*[e3]xpr[e3]ss|\b[a4]m[e3]x\b', 'American Express'),

    # ── CABAL ────────────────────────────────────────────────────
    (r'\bc[a4]b[a4]l\b',             'Cabal'),

    # ── BENEFÍCIOS ───────────────────────────────────────────────
    (r'\b[a4]l[e3]lo\b',             'Alelo'),
    (r'\bt[il1]ck[e3]t\b',          'Ticket'),
    (r'\bsod[e3]xo\b',              'Sodexo'),
    (r'vr\s*(ben[e3]f[il1]c|r[e3]f[e3][il1]|[a4]l[il1]m)', 'VR Benefícios'),
    (r'\bb[e3]n[e3]fl[e3]x\b',      'Beneflex'),

    # ── OUTROS ───────────────────────────────────────────────────
    (r'\b[a4]ur[a4]\b',             'Aura'),
    (r'd[il1]n[e3]rs\s*club|\bd[il1]n[e3]rs\b', 'Diners Club'),
    (r'\bsorocr[e3]d\b',            'Sorocred'),

    # Sicoob aparece em comprovantes PIX
    (r'\bs[il1]co+b\b',             'PIX'),
]

# ── TIPOS DE PAGAMENTO ─────────────────────────────────────────
# Padrão Cielo: "DEBITO A VISTA", "CREDITO A VISTA", "PAGAMENTO PIX"
# Tolerante a erros OCR comuns (O↔0, I↔1, S↔5, etc.)
TIPOS_REGRAS = [
    # PIX — antes de débito para não confundir
    (r'pagamento\s*p[il1]x|p[il1]x\s*r\$|\bp[il1]x\b', 'PIX'),

    # Débito à Vista — inclusive quando OCR cola palavras: "debitoavista"
    (r'd[e3][b8][il1]t[o0].{0,8}[a4@].{0,4}v[il1][s5]t[a4]', 'Débito à Vista'),
    (r'\bdeb[i1l]t[o0]?\s*[a4]?\s*v[i1l]st[a4]\b',        'Débito à Vista'),
    (r'\bdeb[i1l]t[o0]?\b',                                 'Débito à Vista'),

    # Crédito à Vista — inclusive quando OCR cola palavras: "creditoavista"
    (r'cr[e3]d[i1l]t[o0].{0,8}[a4@].{0,4}v[il1][s5]t[a4]', 'Crédito à Vista'),
    (r'\bcred[i1l]t[o0]?\s*[a4]?\s*v[i1l]st[a4]\b',      'Crédito à Vista'),
    (r'\bcred[i1l]t[o0]?\b',                               'Crédito à Vista'),

    # Parcelado
    (r'p[a4]rc[e3]l[a4]do',                       'Parcelado'),

    # Contactless
    (r'cont[a4]ctl[e3]ss|s[e3]m\s*cont[a4]to',    'Contactless'),
]


# ══════════════════════════════════════════════════════════════
#  FUNÇÕES OCR
# ══════════════════════════════════════════════════════════════
def get_reader():
    global _reader
    if _reader is None:
        _reader = easyocr.Reader(['pt', 'en'], gpu=False)
    return _reader


def _normalizar_texto(texto):
    t = (texto or '').lower()
    t = ''.join(
        ch for ch in unicodedata.normalize('NFD', t)
        if unicodedata.category(ch) != 'Mn'
    )
    t = re.sub(r'[^a-z0-9$]+', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t


def _rotacionar(frame, angulo):
    h, w = frame.shape[:2]
    matriz = cv2.getRotationMatrix2D((w // 2, h // 2), angulo, 1.0)
    return cv2.warpAffine(
        frame,
        matriz,
        (w, h),
        flags=cv2.INTER_CUBIC,
        borderMode=cv2.BORDER_REPLICATE,
    )


def extrair_texto(frame, agressivo=False, manter_linhas=False):
    """Extrai texto bruto do frame. Retorna string com todos os tokens."""
    if OCR_ENGINE == "easyocr":
        candidatos = [frame]
        if agressivo:
            candidatos.extend([_rotacionar(frame, -6), _rotacionar(frame, 6)])
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
            gray = clahe.apply(gray)
            bw = cv2.adaptiveThreshold(
                gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                cv2.THRESH_BINARY, 21, 10
            )
            candidatos.append(cv2.cvtColor(bw, cv2.COLOR_GRAY2BGR))

        reader = get_reader()
        textos = []
        for img in candidatos:
            rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            h, w = rgb.shape[:2]
            if w < 780:
                scale = 780 / w
                rgb = cv2.resize(
                    rgb,
                    (int(w * scale), int(h * scale)),
                    interpolation=cv2.INTER_CUBIC,
                )
            res = reader.readtext(rgb, detail=0, paragraph=not manter_linhas)
            if res:
                textos.append('\n'.join(res) if manter_linhas else ' '.join(res))
            if not agressivo and textos:
                break

        return ' '.join(textos)

    elif OCR_ENGINE == "tesseract":
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        # Upscale antes do Tesseract
        gray = cv2.resize(gray, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        gray = clahe.apply(gray)
        binary = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                        cv2.THRESH_BINARY, 21, 10)
        return pytesseract.image_to_string(binary, config='--psm 6 -l por+eng')
    return ""


def _match_regras(texto, regras):
    """Testa lista de (regex, nome) contra texto. Retorna primeiro match."""
    tl = texto.lower()
    tn = _normalizar_texto(texto)
    for padrao, nome in regras:
        if re.search(padrao, tl) or re.search(padrao, tn):
            return nome
    return None


def extrair_bandeira(texto):
    return _match_regras(texto, BANDEIRAS_REGRAS)


def extrair_tipo(texto):
    normalizado = _normalizar_texto(texto)

    # Linha principal da Cielo costuma vir como:
    # "PREPAGO MASTERCARD - DEBITO A VISTA"
    linha_topo = re.search(r'\b(?:prepago\s+)?(?:visa|mastercard|elo|hipercard|amex|cabal)\b\s*[-:]\s*([a-z0-9\s]{6,40})', normalizado)
    if linha_topo:
        tipo_linha = _match_regras(linha_topo.group(1), TIPOS_REGRAS)
        if tipo_linha:
            return tipo_linha

    return _match_regras(texto, TIPOS_REGRAS)


def extrair_pix_instituicao(texto):
    """Retorna instituição do PIX quando identificável no comprovante."""
    tn = _normalizar_texto(texto)
    if not re.search(r'\bp[il1]x\b', tn):
        return None

    if re.search(r'\bs[il1]co+b\b', tn):
        return 'Sicoob'

    if re.search(r'\bcaixa\b|\bcef\b|caixa\s*economica', tn):
        return 'Caixa'

    return None


def _parse_data_str(data_str):
    s = re.sub(r'\s+', '', data_str)
    m = re.match(r'^(\d{2})[/\-.](\d{2})[/\-.](\d{2,4})$', s)
    if not m:
        return None
    d, mo, a = m.groups()
    try:
        aa = int(a)
        if len(a) == 2:
            aa = 2000 + aa if aa < 50 else 1900 + aa
        return date(aa, int(mo), int(d))
    except Exception:
        return None


def extrair_data_pix_layout(texto_linhas):
    """Para PIX, prioriza data no topo (normalmente acompanhada de hora)."""
    if not texto_linhas:
        return None

    # Ex.: "20-04/26 12:05:16" no topo.
    m = re.search(
        r'(?im)^\s*(\d{2}\s*[/\-.]\s*\d{2}\s*[/\-.]\s*\d{2,4})\s+\d{1,2}:\d{2}(?::\d{2})?',
        texto_linhas,
    )
    if m:
        d = _parse_data_str(m.group(1))
        if d:
            return d

    # Fallback: primeira data encontrada nas primeiras linhas.
    for ln in texto_linhas.splitlines()[:6]:
        mm = re.search(r'(\d{2}\s*[/\-.]\s*\d{2}\s*[/\-.]\s*\d{2,4})', ln)
        if mm:
            d = _parse_data_str(mm.group(1))
            if d:
                return d
    return None


def extrair_valor_pix_layout(texto_linhas):
    """Para PIX, prioriza valor na linha do PIX ou da palavra VALOR."""
    if not texto_linhas:
        return None

    linhas = [l.strip() for l in texto_linhas.splitlines() if l.strip()]

    def _parse_valor_str(raw):
        s = raw.strip().replace(' ', '')
        if ',' in s and '.' in s:
            s = s.replace('.', '').replace(',', '.')
        elif ',' in s:
            s = s.replace(',', '.')
        else:
            partes = s.split('.')
            if len(partes) > 2:
                s = ''.join(partes[:-1]) + '.' + partes[-1]
        return float(s)

    # 1) Linha com PIX + valor (comum em Sicoob: "Pix R$ 13.04")
    for ln in linhas:
        if re.search(r'\bp[il1]x\b', ln, re.IGNORECASE):
            # Prioriza número imediatamente após PIX/R$.
            m = re.search(
                r'\bp[il1]x\b[^0-9]{0,10}(?:r\$\s*)?(\d{1,3}(?:[\.,]\d{3})*[\.,]\d{2}|\d+[\.,]\d{2}|\d+\s\d{2})',
                ln,
                re.IGNORECASE,
            )
            if m:
                try:
                    v = _parse_valor_str(m.group(1).replace(' ', '.'))
                    if 0.01 <= v <= 99999.99:
                        return v
                except Exception:
                    pass
            v = extrair_valor(ln)
            if v is not None:
                return v

    # 2) Linha VALOR + valor (comum em Caixa/Cielo PIX)
    for ln in linhas:
        if re.search(r'\bvalor\b', ln, re.IGNORECASE):
            v = extrair_valor(ln)
            if v is not None:
                return v

    return None


def extrair_valor(texto):
    """
    Extrai valor monetário. Prioriza padrão Cielo:
    valor aparece como "R$ 130,01" ou "50,00" (sem R$) perto de "VALOR".
    """
    candidatos = []

    def _parse_valor(raw):
        s = raw.strip().replace(' ', '')
        if ',' in s and '.' in s:
            # Ex.: 1.234,56
            s = s.replace('.', '').replace(',', '.')
        elif ',' in s:
            # Ex.: 14,00
            s = s.replace(',', '.')
        else:
            # Ex.: 14.00
            partes = s.split('.')
            if len(partes) > 2:
                s = ''.join(partes[:-1]) + '.' + partes[-1]
        return float(s)

    def _parece_fragmento_data(txt, ini, fim):
        janela = txt[max(0, ini - 14):min(len(txt), fim + 14)]
        antes = txt[max(0, ini - 4):ini]
        depois = txt[fim:min(len(txt), fim + 4)]

        # Evita capturar parte de data/hora, ex: 20-04/26 12:05
        if re.search(r'\d{1,2}\s*[/\-.]\s*\d{1,2}\s*[/\-.]\s*\d{2,4}', janela):
            return True
        if re.search(r'\d{1,2}:\d{2}(?::\d{2})?', janela):
            return True
        if re.search(r'[-/.]\s*$', antes) or re.search(r'^\s*[-/.]', depois):
            return True
        return False

    # Casos mais confiáveis primeiro (PIX e VALOR explícito)
    for p in [
        r'\bp[il1]x\b\s*(?:r\$\s*)?(\d{1,3}(?:[\.,]\d{3})*[\.,]\d{2}|\d+[\.,]\d{2})',
        r'\br\$\s*(\d{1,3}(?:[\.,]\d{3})*[\.,]\d{2}|\d+[\.,]\d{2})',
        r'\bvalor\b[\s:.-]*(?:r\$\s*)?(\d{1,3}(?:[\.,]\d{3})*[\.,]\d{2}|\d+[\.,]\d{2})',
    ]:
        for m in re.finditer(p, texto, re.IGNORECASE):
            try:
                if _parece_fragmento_data(texto, m.start(1), m.end(1)):
                    continue
                v = _parse_valor(m.group(1))
                if 0.01 <= v <= 99999.99:
                    candidatos.append(v)
            except:
                pass

    # Último recurso: número monetário isolado, com filtro mais forte contra data/hora.
    for m in re.finditer(r'(?<![\d/\-])(\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}|\d+\.\d{2})(?![\d/])', texto, re.IGNORECASE):
        try:
            if _parece_fragmento_data(texto, m.start(1), m.end(1)):
                continue
            v = _parse_valor(m.group(1))
            if 0.01 <= v <= 99999.99:
                candidatos.append(v)
        except:
            pass

    if not candidatos:
        return None
    # Retorna o maior valor encontrado (geralmente o total é o maior)
    return max(candidatos)


def extrair_data(texto):
    """
    Extrai data. Padrão Cielo: DD/MM/AA (ex: 11/04/26).
    Também suporta DD/MM/AAAA e AAAA/MM/DD.
    """
    for p in [
        r'(\d{2})\s*[/\-\.]\s*(\d{2})\s*[/\-\.]\s*(\d{4})',   # DD/MM/YYYY
        r'(\d{4})\s*[/\-\.]\s*(\d{2})\s*[/\-\.]\s*(\d{2})',   # YYYY/MM/DD
        r'(\d{2})\s*[/\-\.]\s*(\d{2})\s*[/\-\.]\s*(\d{2})\b', # DD/MM/YY
    ]:
        m = re.search(p, texto)
        if m:
            try:
                g = m.groups()
                if len(g[2]) == 4:
                    return date(int(g[2]), int(g[1]), int(g[0]))
                elif len(g[0]) == 4:
                    return date(int(g[0]), int(g[1]), int(g[2]))
                else:
                    # DD/MM/YY — assume século 2000 se YY < 50
                    a = 2000 + int(g[2]) if int(g[2]) < 50 else 1900 + int(g[2])
                    return date(a, int(g[1]), int(g[0]))
            except:
                pass
    return None

def fmt_brl(valor):
    return f"R$ {valor:,.2f}".replace(',','X').replace('.',',').replace('X','.')

# ══════════════════════════════════════════════════════════════
#  EXPORTAÇÃO
# ══════════════════════════════════════════════════════════════
def exportar_excel(registros, data_sessao_str, caminho):
    from collections import defaultdict
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Notas {data_sessao_str}"

    def cell(r, c, v, bold=False, color="E8EAF0", bg=None, fmt=None, align='center'):
        cel = ws.cell(row=r, column=c, value=v)
        cel.font = XLFont(color=color, bold=bold, size=10)
        if bg: cel.fill = PatternFill("solid", fgColor=bg)
        if fmt: cel.number_format = fmt
        cel.alignment = Alignment(horizontal=align, vertical='center')
        return cel

    headers = ['#','Bandeira','Tipo','Valor (R$)','Data Nota','Data Sessão','Status','Hora']
    for c, h in enumerate(headers, 1):
        cell(1, c, h, bold=True, color="4F8EF7", bg="111318")
    ws.row_dimensions[1].height = 24

    total = 0.0
    for i, reg in enumerate(registros, 1):
        r = i + 1
        d = reg['data'].strftime('%d/%m/%Y') if reg['data'] else '—'
        st = reg.get('status_data', 'OK')
        bg = "0A1A0F" if st == 'OK' else "1A1500"
        if i % 2 == 0 and st == 'OK': bg = "161A21"
        vals = [i, reg['bandeira'], reg['tipo'],
                reg['valor'] or 0, d, data_sessao_str, st, reg.get('hora','')]
        for c, v in enumerate(vals, 1):
            fmt = 'R$ #,##0.00' if c == 4 else None
            cell(r, c, v, bg=bg, fmt=fmt)
        if reg['valor']: total += reg['valor']

    # Subtotais
    rs = len(registros) + 3
    ws.cell(row=rs, column=1, value='SUBTOTAIS').font = XLFont(bold=True, color="4F8EF7")
    rs += 1
    for h, col in [('Bandeira',1),('Tipo',2),('Qtd',3),('Total',4)]:
        cell(rs, col, h, bold=True, color="9CA3AF")
    rs += 1

    sub = defaultdict(lambda: {'qtd':0,'valor':0.0})
    for reg in registros:
        k = (reg['bandeira'], reg['tipo'])
        sub[k]['qtd'] += 1
        sub[k]['valor'] += reg['valor'] or 0

    for i, ((b, t), d) in enumerate(sorted(sub.items())):
        bg = "161A21" if i%2==0 else "1C1F26"
        cell(rs,1,b,bg=bg); cell(rs,2,t,bg=bg)
        cell(rs,3,d['qtd'],bg=bg)
        cell(rs,4,d['valor'],bg=bg,fmt='R$ #,##0.00')
        rs += 1

    rs += 1
    cell(rs,2,'TOTAL GERAL',bold=True,color="38D9A9",bg="0E1420")
    cell(rs,3,len(registros),bold=True,color="38D9A9",bg="0E1420")
    cell(rs,4,total,bold=True,color="38D9A9",bg="0E1420",fmt='R$ #,##0.00')

    for col, w in zip(range(1,9),[5,22,16,14,14,14,16,10]):
        ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width = w

    wb.save(caminho)
    return total

# ══════════════════════════════════════════════════════════════
#  APP
# ══════════════════════════════════════════════════════════════
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Leitor de Notas · OCR")
        self.root.geometry("1280x820")
        self.root.configure(bg=C['bg'])
        self.root.resizable(True, True)
        self.root.minsize(1100, 720)

        self.cap = None
        self.capturando = False
        self.ocr_ativo = False
        self.frame_atual = None
        self.lock = threading.Lock()
        self.ultimo_ocr = 0
        self.cooldown = 3.0
        self.ocr_em_processamento = False
        self.reader_pronto = OCR_ENGINE != "easyocr"
        self.reader_precarregando = False
        self.preview_width = 500
        self.preview_height = 320
        self.data_sessao = date.today()
        self.registros = []
        self.subtotais = {}   # {(bandeira, tipo): {qtd, valor}}
        self.pendente = None

        self._build()
        self.root.bind('<Return>', self._aceitar)
        self.root.bind('<BackSpace>', self._rejeitar)

    # ────────────────────────────────────────── UI BUILD ──────
    def _build(self):
        # BODY
        body = tk.Frame(self.root, bg=C['bg'])
        body.pack(fill='both', expand=True)

        # COLUNA ESQUERDA
        left = tk.Frame(body, bg=C['bg'], width=self.preview_width + 34)
        left.pack(side='left', fill='y', padx=12, pady=8)
        left.pack_propagate(False)

        self._bloco_camera(left)
        self._bloco_campos(left)
        self._bloco_controles(left)

        # DIVISOR
        tk.Frame(body, bg=C['border'], width=1).pack(side='left', fill='y', pady=12)

        # COLUNA DIREITA
        right = tk.Frame(body, bg=C['bg'])
        right.pack(side='right', fill='both', expand=True, padx=12, pady=8)
        self._bloco_tabela(right)

    # ── BLOCO CÂMERA ──────────────────────────────────────────
    def _bloco_camera(self, parent):
        card = self._card(parent)
        card.pack(fill='x', pady=(0, 8))

        # header
        hdr = tk.Frame(card, bg=C['surface'])
        hdr.pack(fill='x', padx=14, pady=(10, 6))
        self._lbl(hdr, 'Câmera ao vivo', 11, bold=True).pack(side='left')
        self.badge = tk.Label(hdr, text=' INATIVA ',
                               font=('Segoe UI', 8, 'bold'),
                               fg=C['muted'], bg=C['surface2'], padx=6, pady=2)
        self.badge.pack(side='right')

        # preview
        cam_wrap = tk.Frame(
            card,
            bg=C['border'],
            width=self.preview_width + 2,
            height=self.preview_height + 2,
        )
        cam_wrap.pack(fill='x', padx=14, pady=(0, 8))
        cam_wrap.pack_propagate(False)
        self.lbl_cam = tk.Label(cam_wrap, bg='#0A0C10',
                                 text='sem sinal', font=('Segoe UI', 10),
                                 fg=C['muted'])
        self.lbl_cam.pack(fill='both', expand=True, padx=1, pady=1)

    # ── BLOCO CAMPOS DETECTADOS ────────────────────────────────
    def _bloco_campos(self, parent):
        card = self._card(parent)
        card.pack(fill='x', pady=(0, 8))

        hdr = tk.Frame(card, bg=C['surface'])
        hdr.pack(fill='x', padx=14, pady=(10, 4))
        self._lbl(hdr, 'Campos identificados', 11, bold=True).pack(side='left')
        self.lbl_hint = self._lbl(hdr, 'aguardando…', 8, color=C['muted'])
        self.lbl_hint.pack(side='right')

        self._sep(card)

        grid = tk.Frame(card, bg=C['surface'])
        grid.pack(fill='x', padx=16, pady=10)
        grid.columnconfigure(0, weight=1, minsize=96)
        grid.columnconfigure(1, weight=3)

        self.f_bandeira = self._campo_row(grid, 0, 'BANDEIRA')
        self.f_tipo     = self._campo_row(grid, 1, 'TIPO')
        self.f_valor    = self._campo_row(grid, 2, 'VALOR')
        self.f_data     = self._campo_row(grid, 3, 'DATA DA NOTA')

        self._sep(card)
        self.lbl_hint_keys = self._lbl(
            card,
            'Enter: aceitar    Backspace: rejeitar',
            8,
            color=C['muted2']
        )
        self.lbl_hint_keys.pack(anchor='w', padx=14, pady=(8, 10))

    def _campo_row(self, parent, row, nome):
        self._lbl(parent, nome, 8, color=C['muted']).grid(
            row=row, column=0, sticky='w', pady=3, padx=(0, 12))
        lbl = tk.Label(parent, text='—', font=('Segoe UI', 12, 'bold'),
                       fg=C['surface2'], bg=C['surface'], anchor='w')
        lbl.grid(row=row, column=1, sticky='ew', pady=3)
        return lbl

    # ── BLOCO CONTROLES ────────────────────────────────────────
    def _bloco_controles(self, parent):
        card = self._card(parent)
        card.pack(fill='x', pady=(0, 6), side='bottom')

        # Botões câmera
        br = tk.Frame(card, bg=C['surface'])
        br.pack(fill='x', padx=14, pady=(10, 6))
        br.columnconfigure(0, weight=1)
        br.columnconfigure(1, weight=1)

        self.btn_cam = self._btn(br, '▶  Iniciar camera',
                                  bg=C['accent'], fg=C['bg'],
                                  cmd=self._toggle_cam,
                                  bold=True, size=10, padx=12, pady=7)
        self.btn_cam.grid(row=0, column=0, sticky='ew')

        self.btn_ocr_toggle = self._btn(br, '◎  Iniciar leitura',
                                         bg=C['surface2'], fg=C['muted'],
                                         cmd=self._toggle_ocr,
                                         bold=True, size=10, padx=12, pady=7)
        self.btn_ocr_toggle.config(state='disabled')
        self.btn_ocr_toggle.grid(row=0, column=1, sticky='ew', padx=(8, 0))

        # Intervalo
        ir = tk.Frame(card, bg=C['surface'])
        ir.pack(fill='x', padx=14, pady=(0, 6))
        self._lbl(ir, 'Intervalo:', 8, color=C['muted']).pack(side='left')
        self.lbl_int = self._lbl(ir, '3.0s', 8, color=C['accent2'])
        self.lbl_int.pack(side='right')
        self.var_interval = tk.DoubleVar(value=3.0)
        tk.Scale(ir, from_=1, to=8, resolution=0.5, orient='horizontal',
                 variable=self.var_interval, bg=C['surface'], fg=C['muted'],
                 troughcolor=C['surface2'], highlightthickness=0,
                 showvalue=False, sliderlength=14,
                 command=lambda v: [setattr(self,'cooldown',float(v)),
                                    self.lbl_int.config(text=f'{float(v):.1f}s')]
                 ).pack(side='left', fill='x', expand=True, padx=6)

        self.lbl_status = self._lbl(card, 'Câmera inativa', 8, color=C['muted'])

    # ── BLOCO TABELA ──────────────────────────────────────────
    def _bloco_tabela(self, parent):
        hdr = tk.Frame(parent, bg=C['bg'])
        hdr.pack(fill='x', pady=(0, 8))
        self._lbl(hdr, 'Totalizador da sessão', 13, bold=True,
                  bg=C['bg']).pack(side='left')

        self._btn(hdr, '✕ Limpar', bg=C['surface'], fg=C['danger'],
                  cmd=self._limpar).pack(side='right')
        self._btn(hdr, '⬇  Exportar Excel', bg=C['accent2'], fg=C['bg'],
                  cmd=self._exportar, bold=True).pack(side='right', padx=(0,8))

        # Treeview
        card = self._card(parent)
        card.pack(fill='both', expand=True)

        style = ttk.Style()
        style.theme_use('clam')
        style.configure('N.Treeview',
                         background=C['surface'], foreground=C['text'],
                         fieldbackground=C['surface'], rowheight=32,
                         font=('Segoe UI', 10), borderwidth=0)
        style.configure('N.Treeview.Heading',
                         background=C['surface2'], foreground=C['muted2'],
                         font=('Segoe UI', 9, 'bold'), relief='flat')
        style.map('N.Treeview',
                  background=[('selected', '#2A2E3A')],
                  foreground=[('selected', C['accent'])])

        cols = ('bandeira', 'tipo', 'qtd', 'valor')
        self.tree = ttk.Treeview(card, columns=cols, show='headings',
                                  style='N.Treeview')

        for col, label_txt, w in [
            ('bandeira','Bandeira',220), ('tipo','Tipo',120),
            ('qtd','Qtd',70), ('valor','Valor Total',160)
        ]:
            self.tree.heading(col, text=label_txt)
            self.tree.column(col, width=w, anchor='center', minwidth=60)

        self.tree.tag_configure('even',  background='#161A21', foreground=C['text'])
        self.tree.tag_configure('odd',   background=C['surface'], foreground=C['text'])
        self.tree.tag_configure('total', background='#0D1320',
                                foreground=C['accent2'],
                                font=('Segoe UI', 10, 'bold'))

        sb = ttk.Scrollbar(card, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side='left', fill='both', expand=True, padx=(12,0), pady=8)
        sb.pack(side='right', fill='y', pady=8)

        # Footer totais
        self._sep(parent)
        footer = tk.Frame(parent, bg=C['surface'],
                          highlightthickness=1, highlightbackground=C['border'])
        footer.pack(fill='x')

        self.lbl_qtd = tk.Label(footer, text='0 notas',
                                 font=('Segoe UI', 11), fg=C['muted2'],
                                 bg=C['surface'])
        self.lbl_qtd.pack(side='left', padx=16, pady=14)

        self._lbl(footer, 'Total geral:', 9, color=C['muted'],
                  bg=C['surface']).pack(side='right', padx=(0,4), pady=14)

        self.lbl_total = tk.Label(footer, text='R$ 0,00',
                                   font=('Segoe UI', 17, 'bold'),
                                   fg=C['accent2'], bg=C['surface'])
        self.lbl_total.pack(side='right', padx=(0,16), pady=14)

    # ── HELPERS UI ─────────────────────────────────────────────
    def _card(self, parent):
        return tk.Frame(parent, bg=C['surface'],
                        highlightthickness=1, highlightbackground=C['border'])

    def _sep(self, parent):
        tk.Frame(parent, bg=C['border'], height=1).pack(fill='x', padx=0)

    def _lbl(self, parent, text, size=10, bold=False, color=None, bg=None):
        return tk.Label(parent, text=text,
                        font=('Segoe UI', size, 'bold' if bold else 'normal'),
                        fg=color or C['text'],
                        bg=bg if bg else parent.cget('bg'))

    def _btn(self, parent, text, bg, fg, cmd, bold=False, size=10,
             padx=14, pady=7):
        return tk.Button(parent, text=text,
                          font=('Segoe UI', size, 'bold' if bold else 'normal'),
                          bg=bg, fg=fg, activebackground=bg, activeforeground=fg,
                          relief='flat', cursor='hand2', padx=padx, pady=pady,
                          command=cmd)

    def _precarregar_ocr(self):
        if OCR_ENGINE != "easyocr" or self.reader_pronto or self.reader_precarregando:
            return

        self.reader_precarregando = True
        self.lbl_status.config(text='Preparando OCR…', fg=C['warn'])

        def worker():
            try:
                get_reader()
                self.reader_pronto = True
                self.root.after(0, lambda: self.lbl_status.config(
                    text='OCR pronto para leitura', fg=C['accent2']))
            except Exception as e:
                self.root.after(0, lambda msg=str(e): self.lbl_status.config(
                    text=f'Falha ao inicializar OCR: {msg[:80]}', fg=C['danger']))
            finally:
                self.reader_precarregando = False

        threading.Thread(target=worker, daemon=True).start()

    # ── CÂMERA ─────────────────────────────────────────────────
    def _toggle_cam(self):
        if self.capturando: self._parar_cam()
        else: self._iniciar_cam()

    def _iniciar_cam(self):
        self.cap = cv2.VideoCapture(0)
        if not self.cap.isOpened():
            messagebox.showerror('Erro', 'Câmera não encontrada.\nVerifique a conexão.')
            return
        self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        self.capturando = True
        self.btn_cam.config(text='■  Parar câmera', bg=C['danger'],
                             activebackground=C['danger'])
        self.btn_ocr_toggle.config(state='normal', fg=C['text'])
        self.badge.config(text=' ATIVA ', fg=C['accent2'], bg='#0D2E22')
        self.lbl_status.config(text='Câmera ativa — ative a leitura automática',
                                fg=C['muted2'])
        self._precarregar_ocr()
        threading.Thread(target=self._loop_cam, daemon=True).start()

    def _parar_cam(self):
        self.ocr_ativo = False
        self.capturando = False
        self.ocr_em_processamento = False
        if self.cap: self.cap.release()
        self.btn_cam.config(text='▶  Iniciar camera', bg=C['accent'],
                             activebackground=C['accent'])
        self.btn_ocr_toggle.config(state='disabled', fg=C['muted'],
                        text='◎  Iniciar leitura', bg=C['surface2'])
        self.badge.config(text=' INATIVA ', fg=C['muted'], bg=C['surface2'])
        self.lbl_cam.config(image='', text='sem sinal')
        self.lbl_status.config(text='Câmera inativa', fg=C['muted'])

    def _loop_cam(self):
        while self.capturando:
            ret, frame = self.cap.read()
            if not ret: break
            with self.lock:
                self.frame_atual = frame.copy()

            display = frame.copy()
            h, w = display.shape[:2]
            mx, my = int(w*.04), int(h*.04)
            cor = (56,217,169) if self.ocr_ativo else (79,142,247)
            cv2.rectangle(display, (mx,my), (w-mx,h-my), cor, 2)
            txt = 'AGUARDANDO CONFIRMACAO...' if self.pendente else \
                  ('LENDO...' if self.ocr_ativo else 'POSICIONE A NOTA')
            cv2.putText(display, txt, (mx+4, my-8),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.44, cor, 1)

            if self.ocr_ativo and self.pendente is None:
                agora = time.time()
                if (not self.ocr_em_processamento and
                        agora - self.ultimo_ocr >= self.cooldown):
                    self.ultimo_ocr = agora
                    self.ocr_em_processamento = True
                    roi = frame[my:h-my, mx:w-mx].copy()
                    threading.Thread(target=self._processar_wrapper,
                                     args=(roi,), daemon=True).start()

            rgb = cv2.cvtColor(display, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(rgb).resize(
                (self.preview_width, self.preview_height),
                Image.LANCZOS,
            )
            imgtk = ImageTk.PhotoImage(img)
            self.root.after(0, lambda i=imgtk: self._set_cam(i))
            time.sleep(0.033)

    def _set_cam(self, imgtk):
        self.lbl_cam.configure(image=imgtk, text='')
        self.lbl_cam.image = imgtk

    def _toggle_ocr(self):
        self.ocr_ativo = not self.ocr_ativo
        if self.ocr_ativo:
            self._precarregar_ocr()
            self.btn_ocr_toggle.config(text='◉  Pausar leitura',
                                        bg=C['warn'], fg=C['bg'],
                                        activebackground=C['warn'])
            self.lbl_status.config(text='Leitura automática ativa…', fg=C['accent2'])
        else:
            self.btn_ocr_toggle.config(text='◎  Iniciar leitura',
                                        bg=C['surface2'], fg=C['text'],
                                        activebackground=C['surface2'])
            self.lbl_status.config(text='Leitura pausada', fg=C['muted'])

    def _processar_wrapper(self, frame):
        try:
            self._processar(frame)
        finally:
            self.ocr_em_processamento = False

    # ── OCR ────────────────────────────────────────────────────
    def _processar(self, frame):
        self.root.after(0, lambda: self.lbl_status.config(
            text='Processando…', fg=C['warn']))
        texto = extrair_texto(frame)
        bandeira = extrair_bandeira(texto)
        tipo = extrair_tipo(texto)
        valor = extrair_valor(texto)
        data = extrair_data(texto)

        # Fallback mais forte para comprovante inclinado/ruidoso.
        # Só executa quando bandeira ou tipo falham na primeira passada.
        if bandeira is None or tipo is None:
            texto_extra = extrair_texto(frame, agressivo=True)
            if texto_extra:
                texto = f'{texto} {texto_extra}'.strip()
                bandeira = bandeira or extrair_bandeira(texto_extra)
                tipo = tipo or extrair_tipo(texto_extra)
                valor = valor if valor is not None else extrair_valor(texto_extra)
                data = data or extrair_data(texto_extra)

        # Ajuste por layout específico de PIX (campos em posições diferentes).
        texto_pix = ''
        if re.search(r'\bp[il1]x\b', _normalizar_texto(texto)):
            texto_pix = extrair_texto(frame, manter_linhas=True)
            valor_pix = extrair_valor_pix_layout(texto_pix)
            data_pix = extrair_data_pix_layout(texto_pix)
            if valor_pix is not None:
                valor = valor_pix
            if data_pix is not None:
                data = data_pix

        # PIX por instituição:
        # - Bandeira sempre "PIX"
        # - Tipo: "Sicoob" quando identificar; caso contrário "Caixa"
        texto_unificado = f'{texto} {texto_pix}'.strip()
        pix_detectado = (
            re.search(r'\bp[il1]x\b', _normalizar_texto(texto_unificado)) is not None
            or bandeira == 'PIX'
            or tipo == 'PIX'
        )
        if pix_detectado:
            pix_inst = extrair_pix_instituicao(texto_unificado)
            bandeira = 'PIX'
            tipo = pix_inst if pix_inst else 'Caixa'
            if valor is None and texto_pix:
                # Última tentativa específica de PIX por linha.
                valor = extrair_valor_pix_layout(texto_pix)
            if data is None and texto_pix:
                data = extrair_data_pix_layout(texto_pix)

        if valor is None:
            trecho = texto[:80].replace(chr(10), ' ') if texto else '(sem texto)'
            self.root.after(0, lambda t=trecho: self.lbl_status.config(
                text=f'Valor não encontrado. OCR: "{t}"', fg=C['muted']))
            return

        self.root.after(0, self._exibir_campos, {
            'bandeira': bandeira or 'Não identificada',
            'tipo': tipo or 'Não identificado',
            'valor': valor, 'data': data,
            'texto_bruto': texto,
        })

    def _exibir_campos(self, r):
        self.pendente = r
        self.f_bandeira.config(text=r['bandeira'], fg=C['text'])
        self.f_tipo.config(text=r['tipo'], fg=C['text'])
        self.f_valor.config(text=fmt_brl(r['valor']), fg=C['accent2'])

        if r['data']:
            ok = r['data'] == self.data_sessao
            self.f_data.config(text=r['data'].strftime('%d/%m/%Y'),
                                fg=C['success'] if ok else C['warn'])
        else:
            self.f_data.config(text='Não identificada', fg=C['muted'])

        self.lbl_hint.config(text='↵ aceitar   ⌫ rejeitar', fg=C['accent'])
        self.lbl_status.config(text='Nota identificada — confirme ou rejeite',
                                fg=C['accent2'])

    def _aceitar(self, event=None):
        if not self.pendente: return
        r = self.pendente
        data_nota = r['data']

        if data_nota and data_nota != self.data_sessao:
            s = self.data_sessao.strftime('%d/%m/%Y')
            n = data_nota.strftime('%d/%m/%Y')
            if not messagebox.askyesno('Data divergente',
                f'Data da nota ({n}) difere da sessão ({s}).\nAdicionar mesmo assim?'):
                self._limpar_campos()
                return
            r['status_data'] = '⚠ Divergente'
        elif not data_nota:
            r['status_data'] = 'S/data'
        else:
            r['status_data'] = 'OK'

        r['hora'] = datetime.now().strftime('%H:%M:%S')
        self.registros.append(r)
        self._atualizar_tabela(r)
        self._limpar_campos()

    def _rejeitar(self, event=None):
        if not self.pendente: return
        self._limpar_campos()
        self.lbl_status.config(text='Nota rejeitada — aguardando próxima…',
                                fg=C['muted'])

    def _limpar_campos(self):
        self.pendente = None
        for f in (self.f_bandeira, self.f_tipo, self.f_valor, self.f_data):
            f.config(text='—', fg=C['surface2'])
        self.lbl_hint.config(text='aguardando…', fg=C['muted'])
        if self.ocr_ativo:
            self.lbl_status.config(text='Leitura automática ativa…', fg=C['accent2'])

    # ── TABELA SUBTOTAIS ───────────────────────────────────────
    def _atualizar_tabela(self, reg):
        key = (reg['bandeira'], reg['tipo'])
        if key not in self.subtotais:
            self.subtotais[key] = {'qtd': 0, 'valor': 0.0}
        self.subtotais[key]['qtd'] += 1
        self.subtotais[key]['valor'] += reg['valor'] or 0.0

        for item in self.tree.get_children():
            self.tree.delete(item)

        total_qtd = 0
        total_val = 0.0
        for i, ((b, t), d) in enumerate(sorted(self.subtotais.items())):
            self.tree.insert('', 'end',
                              values=(b, t, d['qtd'], fmt_brl(d['valor'])),
                              tags=('even' if i%2==0 else 'odd',))
            total_qtd += d['qtd']
            total_val += d['valor']

        self.tree.insert('', 'end',
                          values=('TOTAL GERAL', '—', total_qtd, fmt_brl(total_val)),
                          tags=('total',))

        self.lbl_qtd.config(text=f'{total_qtd} nota{"s" if total_qtd!=1 else ""}')
        self.lbl_total.config(text=fmt_brl(total_val))

    # ── UTILITÁRIOS ────────────────────────────────────────────
    def _sync_data(self):
        try:
            self.data_sessao = date(int(self.var_a.get()),
                                    int(self.var_m.get()),
                                    int(self.var_d.get()))
            dias = ['Seg','Ter','Qua','Qui','Sex','Sáb','Dom']
            d = self.data_sessao
            self.lbl_dia_nome.config(
                text=f'  {dias[d.weekday()]}, {d.strftime("%d/%m/%Y")}')
        except ValueError:
            pass

    def _tick(self):
        if hasattr(self, 'lbl_clock'):
            self.lbl_clock.config(text=datetime.now().strftime('%H:%M:%S'))
            self.root.after(1000, self._tick)

    def _exportar(self):
        if not self.registros:
            messagebox.showwarning('Vazio', 'Nenhum registro para exportar.')
            return
        nome = (f"notas_{self.data_sessao.strftime('%Y%m%d')}_"
                f"{datetime.now().strftime('%H%M%S')}.xlsx")
        caminho = os.path.join(os.path.expanduser('~'), 'Desktop', nome)
        try:
            total = exportar_excel(self.registros,
                                   self.data_sessao.strftime('%d/%m/%Y'), caminho)
            messagebox.showinfo('Exportado!',
                                f'Arquivo salvo em:\n{caminho}\n\n'
                                f'{len(self.registros)} nota(s) · {fmt_brl(total)}')
        except Exception as e:
            messagebox.showerror('Erro ao exportar', str(e))

    def _limpar(self):
        if self.registros and messagebox.askyesno('Limpar', 'Limpar todos os registros?'):
            self.registros.clear()
            self.subtotais.clear()
            for item in self.tree.get_children():
                self.tree.delete(item)
            self.lbl_qtd.config(text='0 notas')
            self.lbl_total.config(text='R$ 0,00')

    def on_close(self):
        self._parar_cam()
        self.root.destroy()


if __name__ == '__main__':
    root = tk.Tk()
    app = App(root)
    root.protocol('WM_DELETE_WINDOW', app.on_close)
    root.mainloop()
