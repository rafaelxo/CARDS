# ─── excel_manager.py ──────────────────────────────────────────────────────────
# Escreve sessões na planilha mestre (notas_cartao.xlsx).
# Abas mensais: A=# | B=Bandeira | C=Tipo | D=Data | E=Valor | F=Hora | G=Status
# Painel!B1  = data do dia selecionado → TEXT(B1,"YYYY-MM") aponta a aba mensal
# Resumo!B1  = =Painel!B1             → herda automaticamente, mostra o mês
# ──────────────────────────────────────────────────────────────────────────────

from __future__ import annotations
import os
from datetime import date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Paleta ───────────────────────────────────────────────────────────────────
W   = 'FFFFFFFF'
G1  = 'FFF1F3F5'
G2  = 'FFE9ECEF'
G4  = 'FF6C757D'
G5  = 'FF212529'
NAV = 'FF1B2A3B'
BL  = 'FF1971C2'
BL2 = 'FFD0EBFF'
BL3 = 'FFE7F5FF'
GR  = 'FF2F9E44'
GR2 = 'FFD3F9D8'

MESES_PT = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
            'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
DIAS_PT  = ['Seg','Ter','Qua','Qui','Sex','Sáb','Dom']
BRL_FMT  = 'R$ #,##0.00'

def fmt_brl(v: float) -> str:
    return f"R$ {v:,.2f}".replace(',','X').replace('.',',').replace('X','.')

def _fl(h):
    h = h.lstrip('#')
    return PatternFill('solid', fgColor=('FF'+h) if len(h)==6 else h)

def _fn(size=10, bold=False, col=G5):
    return Font(name='Calibri', size=size, bold=bold, color=col)

def _al(h='center', v='center'):
    return Alignment(horizontal=h, vertical=v)

def _bd(color=G2):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _c(ws, r, col, val=None, *, bold=False, size=10, fg=G5, bg=W, fmt=None, h='center'):
    c = ws.cell(r, col, val)
    c.font      = _fn(size, bold, fg)
    c.fill      = _fl(bg)
    c.alignment = _al(h)
    c.border    = _bd()
    if fmt: c.number_format = fmt
    return c

def _nome_aba(d: date) -> str:
    """Formato YYYY-MM — espelho de TEXT(B1,"YYYY-MM") no Painel/Resumo."""
    return d.strftime('%Y-%m')

# ── Criar aba mensal de dados ─────────────────────────────────────────────────
def _criar_aba(wb, nome: str, ano: int, mes: int):
    ws = wb.create_sheet(nome)
    ws.sheet_view.showGridLines = False
    ws.tab_color = '1971C2'

    for col, w in zip(range(1, 8), [5, 22, 18, 13, 15, 10, 13]):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Banner
    ws.merge_cells('A1:G1')
    ws.row_dimensions[1].height = 34
    c = ws['A1']
    c.value     = f'Notas  —  {MESES_PT[mes-1]} {ano}'
    c.font      = Font(name='Calibri', size=13, bold=True, color=NAV)
    c.fill      = _fl(W)
    c.alignment = _al('left')

    # Linha azul decorativa
    ws.merge_cells('A2:G2')
    ws['A2'].fill = _fl(BL)
    ws.row_dimensions[2].height = 3

    # Total mensal no topo
    ws.row_dimensions[3].height = 8
    ws.row_dimensions[4].height = 28
    ws.merge_cells('A4:C4')
    c = ws['A4']
    c.value     = 'TOTAL DO MÊS'
    c.font      = _fn(11, True, W)
    c.fill      = _fl(NAV)
    c.alignment = _al('left')
    c.border    = _bd(NAV)
    for col in [4, 6, 7]:
        ws.cell(4, col).fill   = _fl(NAV)
        ws.cell(4, col).border = _bd(NAV)
    e4 = ws['E4']
    e4.value         = '=IFERROR(SUM(E8:E9999),0)'
    e4.font          = _fn(13, True, BL2)
    e4.fill          = _fl(NAV)
    e4.alignment     = _al('center')
    e4.border        = _bd(NAV)
    e4.number_format = BRL_FMT

    ws.row_dimensions[5].height = 8

    # Data validations
    dv_b = DataValidation(type='list', formula1='ListaBandeiras',
                          allow_blank=True, showDropDown=False,
                          showErrorMessage=True, errorTitle='Inválido',
                          error='Use a lista.')
    dv_t = DataValidation(type='list', formula1='ListaTipos',
                          allow_blank=True, showDropDown=False,
                          showErrorMessage=True, errorTitle='Inválido',
                          error='Use a lista.')
    ws.add_data_validation(dv_b)
    ws.add_data_validation(dv_t)
    dv_b.sqref = 'B8:B9999'
    dv_t.sqref = 'C8:C9999'

    ws.freeze_panes = 'A8'
    return ws, 8

# ── Escrever bloco de um dia ──────────────────────────────────────────────────
def _escrever_dia(ws, row: int, d: date, registros: list[dict]) -> int:
    label = f'  {DIAS_PT[d.weekday()]}  {d.strftime("%d/%m/%Y")}'
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row, 1, label)
    c.font      = _fn(10, True, NAV)
    c.fill      = _fl(BL3)
    c.alignment = _al('left')
    c.border    = Border(
        left=Side(style='medium', color=BL),
        right=Side(style='thin',  color=G2),
        top=Side(style='medium',  color=BL),
        bottom=Side(style='thin', color=G2),
    )
    ws.row_dimensions[row].height = 22
    row += 1

    for col, h in enumerate(['#', 'Bandeira', 'Tipo', 'Data', 'Valor (R$)', 'Hora', 'Status'], 1):
        c = ws.cell(row, col, h)
        c.font      = _fn(9, True, BL)
        c.fill      = _fl(BL2)
        c.alignment = _al('center')
        c.border    = Border(
            left=Side(style='thin',     color=G2),
            right=Side(style='thin',    color=G2),
            top=Side(style='thin',      color=G2),
            bottom=Side(style='medium', color=BL),
        )
    ws.row_dimensions[row].height = 20
    row += 1

    inicio = row
    for j, reg in enumerate(registros, 1):
        bg = G1 if j % 2 == 0 else W
        st = reg.get('status_data', 'OK')
        _c(ws, row, 1, j,                    bg=bg, size=9,  fg=G4)
        _c(ws, row, 2, reg['bandeira'],       bg=bg, size=10, fg=NAV, bold=True, h='left')
        _c(ws, row, 3, reg['tipo'],           bg=bg, size=10, fg=G5)
        _c(ws, row, 4, d,                     bg=bg, fmt='DD/MM/YYYY')
        _c(ws, row, 5, reg['valor'] or 0,     bg=bg, bold=True, fg=NAV, fmt=BRL_FMT)
        _c(ws, row, 6, reg.get('hora', ''),   bg=bg, size=9,  fg=G4)
        cs = _c(ws, row, 7, st,               bg=bg, size=9)
        if st != 'OK':
            cs.font = _fn(9, True, 'FFD97706')
        ws.row_dimensions[row].height = 22
        row += 1

    fim = row - 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row, 1, f'  Subtotal  {DIAS_PT[d.weekday()]}  {d.strftime("%d/%m")}')
    c.font = _fn(9, True, GR); c.fill = _fl(GR2)
    c.alignment = _al('left'); c.border = _bd(GR)
    ws.row_dimensions[row].height = 20

    sub = ws.cell(row, 5, f'=SUM(E{inicio}:E{fim})')
    sub.font = _fn(10, True, GR); sub.fill = _fl(GR2)
    sub.alignment = _al('center'); sub.border = _bd(GR)
    sub.number_format = BRL_FMT

    for col in [6, 7]:
        ws.cell(row, col).fill   = _fl(GR2)
        ws.cell(row, col).border = _bd(GR)
    row += 2
    return row

# ── ExcelManager ─────────────────────────────────────────────────────────────
class ExcelManager:
    def __init__(self, caminho: str):
        self.caminho = caminho
        self.wb = load_workbook(caminho) if os.path.exists(caminho) else Workbook()

    def anexar_sessao(self, registros: list[dict], d: date) -> float:
        if not registros:
            return 0.0
        nome = _nome_aba(d)
        if nome in self.wb.sheetnames:
            ws = self.wb[nome]
            next_row = max(ws.max_row + 2, 8)
        else:
            ws, next_row = _criar_aba(self.wb, nome, d.year, d.month)
            self._ordenar_abas()
        _escrever_dia(ws, next_row, d, registros)
        self._atualizar_data_painel(d)
        total = sum(r['valor'] or 0 for r in registros)
        self.wb.save(self.caminho)
        return total

    def _atualizar_data_painel(self, d: date) -> None:
        """
        Escreve a data da sessão em Painel!B1.
        As fórmulas do Painel/Resumo usam TEXT(B1,"YYYY-MM") para apontar
        à aba mensal correta. Resumo!B1 = =Painel!B1, então fica sincronizado.
        """
        if 'Painel' not in self.wb.sheetnames:
            return
        c = self.wb['Painel']['B1']
        c.value         = d
        c.number_format = 'DD/MM/YYYY'

    def _ordenar_abas(self):
        prioridade = ['Config', 'Painel', 'Resumo']
        ordem = [s for s in prioridade if s in self.wb.sheetnames]
        ordem += sorted(s for s in self.wb.sheetnames if s not in prioridade)
        for i, nome in enumerate(ordem):
            idx = self.wb.sheetnames.index(nome)
            if idx != i:
                self.wb.move_sheet(nome, offset=i - idx)
